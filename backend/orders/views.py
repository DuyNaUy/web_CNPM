from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from uuid import uuid4
from decimal import Decimal
from .models import Order, OrderItem, Cart, CartItem
from .serializers import OrderSerializer, OrderCreateSerializer, CartSerializer, CartItemDetailSerializer
from products.models import Product, ProductVariant
from .payment_utils import MoMoPayment
import logging
import json

logger = logging.getLogger(__name__)


class CartViewSet(viewsets.ViewSet):
    """ViewSet cho giỏ hàng"""
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def my_cart(self, request):
        """Lấy giỏ hàng của user hiện tại"""
        try:
            cart, created = Cart.objects.get_or_create(user=request.user)
            serializer = CartSerializer(cart, context={'request': request})
            return Response(serializer.data)
        except Exception as e:
            logger.error(f"Error getting cart: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def add_item(self, request):
        """Thêm sản phẩm vào giỏ hàng"""
        try:
            product_id = request.data.get('product_id')
            quantity = int(request.data.get('quantity', 1))
            unit = request.data.get('unit', '')
            
            if not product_id:
                return Response({'error': 'product_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if quantity <= 0:
                return Response({'error': 'quantity must be greater than 0'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Lấy sản phẩm
            try:
                product = Product.objects.get(id=product_id)
            except Product.DoesNotExist:
                return Response({'error': 'Product not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Kiểm tra stock từ variant nếu có unit
            if unit and product.variants.exists():
                variant = product.variants.filter(size=unit).first()
                if not variant:
                    return Response(
                        {'error': f'Kích thước {unit} không tồn tại'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                if variant.stock < quantity:
                    return Response(
                        {'error': f'Số lượng tồn kho không đủ. Tồn kho: {variant.stock}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            else:
                # Kiểm tra số lượng tồn kho từ product.stock
                if product.stock < quantity:
                    return Response(
                        {'error': f'Số lượng tồn kho không đủ. Tồn kho: {product.stock}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Lấy hoặc tạo giỏ hàng
            cart, created = Cart.objects.get_or_create(user=request.user)
            
            # Lấy giá từ variant hoặc product
            item_price = product.price
            if unit and product.variants.exists():
                variant = product.variants.filter(size=unit).first()
                if variant:
                    item_price = variant.price
            
            # Thêm hoặc cập nhật sản phẩm trong giỏ hàng
            cart_item, item_created = CartItem.objects.get_or_create(
                cart=cart,
                product=product,
                unit=unit,
                defaults={'quantity': quantity, 'price': item_price}
            )
            
            if not item_created:
                # Nếu sản phẩm đã có trong giỏ, cập nhật số lượng
                new_quantity = cart_item.quantity + quantity
                
                # Kiểm tra stock của variant mới
                if unit and product.variants.exists():
                    variant = product.variants.filter(size=unit).first()
                    if variant and variant.stock < new_quantity:
                        return Response(
                            {'error': f'Số lượng tồn kho không đủ. Tồn kho: {variant.stock}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    if product.stock < new_quantity:
                        return Response(
                            {'error': f'Số lượng tồn kho không đủ. Tồn kho: {product.stock}'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                cart_item.quantity = new_quantity
                cart_item.save()
            
            # Trả về giỏ hàng được cập nhật
            serializer = CartSerializer(cart, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except ValueError as e:
            return Response({'error': 'Invalid quantity'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error adding item to cart: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def update_item(self, request):
        """Cập nhật số lượng sản phẩm trong giỏ hàng"""
        try:
            item_id = request.data.get('item_id')
            quantity = int(request.data.get('quantity', 1))
            
            if not item_id:
                return Response({'error': 'item_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            if quantity <= 0:
                return Response({'error': 'quantity must be greater than 0'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Lấy giỏ hàng của user
            try:
                cart = Cart.objects.get(user=request.user)
            except Cart.DoesNotExist:
                return Response({'error': 'Cart not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Lấy mục trong giỏ hàng
            try:
                cart_item = CartItem.objects.get(id=item_id, cart=cart)
            except CartItem.DoesNotExist:
                return Response({'error': 'Cart item not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Kiểm tra số lượng tồn kho
            if cart_item.product.stock < quantity:
                return Response(
                    {'error': f'Số lượng tồn kho không đủ. Tồn kho: {cart_item.product.stock}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Cập nhật số lượng
            cart_item.quantity = quantity
            cart_item.save()
            
            # Trả về giỏ hàng được cập nhật
            serializer = CartSerializer(cart, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except ValueError as e:
            return Response({'error': 'Invalid quantity'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error updating cart item: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def remove_item(self, request):
        """Xóa sản phẩm khỏi giỏ hàng"""
        try:
            item_id = request.data.get('item_id')
            
            if not item_id:
                return Response({'error': 'item_id is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Lấy giỏ hàng của user
            try:
                cart = Cart.objects.get(user=request.user)
            except Cart.DoesNotExist:
                return Response({'error': 'Cart not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Xóa mục trong giỏ hàng
            try:
                cart_item = CartItem.objects.get(id=item_id, cart=cart)
                cart_item.delete()
            except CartItem.DoesNotExist:
                return Response({'error': 'Cart item not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Trả về giỏ hàng được cập nhật
            serializer = CartSerializer(cart, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Error removing cart item: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def clear_cart(self, request):
        """Xóa tất cả sản phẩm khỏi giỏ hàng"""
        try:
            # Lấy giỏ hàng của user
            try:
                cart = Cart.objects.get(user=request.user)
                cart.items.all().delete()
            except Cart.DoesNotExist:
                pass
            
            # Tạo giỏ hàng mới rỗng
            cart, created = Cart.objects.get_or_create(user=request.user)
            serializer = CartSerializer(cart, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Error clearing cart: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class OrderViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def my_orders(self, request):
        """Lấy đơn hàng của user hiện tại"""
        orders = Order.objects.filter(user=request.user)
        serializer = OrderSerializer(orders, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def cancel_order(self, request):
        """Khách hàng hủy đơn hàng của mình"""
        try:
            order_id = request.data.get('order_id')
            
            if not order_id:
                return Response(
                    {'error': 'order_id là bắt buộc'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Lấy đơn hàng của user
            try:
                order = Order.objects.get(id=order_id, user=request.user)
            except Order.DoesNotExist:
                return Response(
                    {'error': 'Đơn hàng không tồn tại hoặc bạn không có quyền hủy đơn hàng này'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Kiểm tra trạng thái đơn hàng - chỉ cho phép hủy khi ở pending hoặc confirmed
            if order.status not in ['pending', 'confirmed']:
                status_names = {
                    'shipping': 'Đang giao',
                    'delivered': 'Đã giao',
                    'cancelled': 'Đã hủy'
                }
                return Response(
                    {'error': f'Không thể hủy đơn hàng ở trạng thái {status_names.get(order.status, order.status)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            old_status = order.status
            order.status = 'cancelled'
            order.save()
            
            # Hoàn trả tồn kho khi hủy đơn
            for order_item in order.items.all():
                product = order_item.product
                if product:
                    unit = order_item.unit
                    quantity = order_item.quantity
                    
                    # Nếu có unit (variant), hoàn trả stock cho variant
                    if unit and product.variants.exists():
                        variant = product.variants.filter(size=unit).first()
                        if variant:
                            variant.stock += quantity
                            variant.save()
                            logger.info(f"Restored variant {variant.size} stock: +{quantity}")
                    else:
                        # Hoàn trả stock cho product
                        product.stock += quantity
                        product.save()
                        logger.info(f"Restored product {product.name} stock: +{quantity}")
            
            serializer = OrderSerializer(order)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Exception as e:
            logger.error(f"Error cancelling order: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def create_order(self, request):
        """Tạo đơn hàng mới"""
        serializer = OrderCreateSerializer(data=request.data)
        if serializer.is_valid():
            try:
                # Lấy dữ liệu
                items_data = serializer.validated_data.get('items', [])
                
                if not items_data:
                    return Response(
                        {'error': 'Đơn hàng phải có ít nhất 1 sản phẩm'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Tính tổng tiền và validate stock
                subtotal = Decimal(0)
                order_items = []
                
                for item_data in items_data:
                    try:
                        product_id = int(item_data.get('id'))
                        quantity = int(item_data.get('quantity', 1))
                        unit = item_data.get('unit', '')
                        
                        # IMPORTANT: Use price from item_data if provided, otherwise use product.price
                        item_price = Decimal(str(item_data.get('price', 0)))
                        if item_price == 0:
                            # Fallback to product price if not provided in item_data
                            product = Product.objects.get(id=product_id)
                            item_price = Decimal(str(product.price))
                        else:
                            product = Product.objects.get(id=product_id)
                        
                        # Validate stock trước khi tạo order
                        if unit and product.variants.exists():
                            variant = product.variants.filter(size=unit).first()
                            if not variant:
                                return Response(
                                    {'error': f'Kích thước {unit} của sản phẩm {product.name} không tồn tại'},
                                    status=status.HTTP_400_BAD_REQUEST
                                )
                            if variant.stock < quantity:
                                return Response(
                                    {'error': f'Sản phẩm {product.name} (Size {unit}) không đủ hàng. Còn lại: {variant.stock}'},
                                    status=status.HTTP_400_BAD_REQUEST
                                )
                        else:
                            if product.stock < quantity:
                                return Response(
                                    {'error': f'Sản phẩm {product.name} không đủ hàng. Còn lại: {product.stock}'},
                                    status=status.HTTP_400_BAD_REQUEST
                                )
                        
                        subtotal += item_price * quantity
                        order_items.append({
                            'product': product,
                            'product_name': product.name,
                            'product_price': item_price,
                            'quantity': quantity,
                            'unit': unit or product.unit or '',
                        })
                    except (ValueError, Product.DoesNotExist):
                        return Response(
                            {'error': f'Sản phẩm không tồn tại'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                
                # Tính phí vận chuyển
                shipping_fee = Decimal(0) if subtotal >= Decimal(500000) else Decimal(30000)
                total_amount = subtotal + shipping_fee
                
                # Tạo order
                order = Order.objects.create(
                    user=request.user,
                    order_code='',  # Sẽ được cập nhật sau khi có ID
                    status='pending',
                    full_name=serializer.validated_data['full_name'],
                    phone=serializer.validated_data['phone'],
                    email=serializer.validated_data['email'],
                    address=serializer.validated_data['address'],
                    city=serializer.validated_data['city'],
                    district=serializer.validated_data['district'],
                    note=serializer.validated_data.get('note', ''),
                    payment_method=serializer.validated_data['payment_method'],
                    payment_status='completed',  # Tạm để completed vì chưa tích hợp payment gateway
                    subtotal=subtotal,
                    shipping_fee=shipping_fee,
                    total_amount=total_amount,
                )
                
                # Cập nhật order_code với format DH + 3 chữ số ID
                order.order_code = f"DH{order.id:03d}"
                order.save(update_fields=['order_code'])
                
                # Tạo order items và cập nhật stock
                for item in order_items:
                    OrderItem.objects.create(
                        order=order,
                        product=item['product'],
                        product_name=item['product_name'],
                        product_price=item['product_price'],
                        quantity=item['quantity'],
                        unit=item['unit'],
                    )
                    
                    # Cập nhật tồn kho
                    product = item['product']
                    unit = item['unit']
                    quantity = item['quantity']
                    
                    # Nếu có unit (variant), cập nhật stock của variant
                    if unit and product.variants.exists():
                        variant = product.variants.filter(size=unit).first()
                        if variant:
                            variant.stock -= quantity
                            variant.save()
                            logger.info(f"Updated variant {variant.size} stock: {variant.stock + quantity} -> {variant.stock}")
                    else:
                        # Cập nhật stock của product
                        product.stock -= quantity
                        product.save()
                        logger.info(f"Updated product {product.name} stock: {product.stock + quantity} -> {product.stock}")
                    
                    # Lưu ý: sold_count sẽ được cập nhật khi order được delivered, không phải lúc tạo
                
                # Xử lý thanh toán MoMo
                if order.payment_method == 'momo':
                    try:
                        # Tạo payment URL từ MoMo
                        momo_response = MoMoPayment.create_payment(
                            order_id=order.id,
                            amount=int(total_amount),
                            order_info=f"Thanh toan don hang {order.order_code}"
                        )
                        
                        if momo_response.get('resultCode') == 0:
                            # Lưu thông tin MoMo
                            order.momo_request_id = momo_response.get('requestId')
                            order.momo_order_id = momo_response.get('orderId')
                            order.save(update_fields=['momo_request_id', 'momo_order_id'])
                            
                            # Trả về payUrl để redirect
                            response_data = OrderSerializer(order).data
                            response_data['payUrl'] = momo_response.get('payUrl')
                            response_data['deeplink'] = momo_response.get('deeplink')
                            response_data['qrCodeUrl'] = momo_response.get('qrCodeUrl')
                            
                            return Response(response_data, status=status.HTTP_201_CREATED)
                        else:
                            # MoMo trả về lỗi
                            return Response(
                                {
                                    'error': f"Lỗi tạo thanh toán MoMo: {momo_response.get('message')}",
                                    'order': OrderSerializer(order).data
                                },
                                status=status.HTTP_400_BAD_REQUEST
                            )
                    except Exception as e:
                        logger.error(f"Error creating MoMo payment: {str(e)}")
                        return Response(
                            {
                                'error': f'Lỗi kết nối MoMo: {str(e)}',
                                'order': OrderSerializer(order).data
                            },
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR
                        )
                
                response_serializer = OrderSerializer(order)
                return Response(response_serializer.data, status=status.HTTP_201_CREATED)
            
            except Exception as e:
                logger.error(f"Error creating order: {str(e)}")
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def all_orders(self, request):
        """Admin - Lấy tất cả đơn hàng"""
        # Check if user is admin (has is_staff or role='admin')
        if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
            return Response(
                {'error': 'Bạn không có quyền truy cập'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Tìm kiếm đơn hàng
        search_query = request.query_params.get('search', None)
        orders = Order.objects.all()
        
        if search_query:
            from django.db.models import Q
            orders = orders.filter(
                Q(order_code__icontains=search_query) |
                Q(full_name__icontains=search_query) |
                Q(phone__icontains=search_query) |
                Q(email__icontains=search_query)
            )
        
        orders = orders.order_by('-created_at')
        serializer = OrderSerializer(orders, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def update_order_status(self, request):
        """Admin - Cập nhật trạng thái đơn hàng"""
        # Check if user is admin
        if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
            return Response(
                {'error': 'Bạn không có quyền truy cập'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            order_id = request.data.get('order_id')
            new_status = request.data.get('status')
            
            if not order_id or not new_status:
                return Response(
                    {'error': 'order_id và status là bắt buộc'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Kiểm tra trạng thái hợp lệ
            valid_statuses = ['pending', 'confirmed', 'shipping', 'delivered', 'cancelled']
            if new_status not in valid_statuses:
                return Response(
                    {'error': f'Trạng thái không hợp lệ. Chọn từ: {", ".join(valid_statuses)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            order = Order.objects.get(id=order_id)
            old_status = order.status
            
            # Định nghĩa thứ tự trạng thái cho admin
            status_flow = {
                'pending': ['confirmed', 'cancelled'],
                'confirmed': ['shipping', 'cancelled'],
                'shipping': ['delivered'],  # Đang giao không thể hủy
                'delivered': [],  # Không thể thay đổi từ delivered
                'cancelled': []   # Không thể thay đổi từ cancelled
            }
            
            # Kiểm tra xem có thể chuyển sang trạng thái mới không
            if old_status == new_status:
                return Response(
                    {'error': f'Đơn hàng đã ở trạng thái {order.get_status_display()}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            allowed_statuses = status_flow.get(old_status, [])
            if new_status not in allowed_statuses:
                status_names = {
                    'pending': 'Chờ xử lý',
                    'confirmed': 'Đã xác nhận',
                    'shipping': 'Đang giao',
                    'delivered': 'Đã giao',
                    'cancelled': 'Đã hủy'
                }
                
                if not allowed_statuses:
                    return Response(
                        {'error': f'Không thể thay đổi trạng thái từ {status_names[old_status]}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                allowed_names = [status_names[s] for s in allowed_statuses]
                return Response(
                    {'error': f'Từ trạng thái {status_names[old_status]}, chỉ có thể chuyển sang: {", ".join(allowed_names)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            order.status = new_status
            order.save()
            
            # Cập nhật sold_count khi order được delivered
            if new_status == 'delivered' and old_status != 'delivered':
                # Lấy tất cả items trong order
                for order_item in order.items.all():
                    product = order_item.product
                    if product:
                        product.sold_count += order_item.quantity
                        product.save(update_fields=['sold_count'])
                        logger.info(f"Updated product {product.name} sold_count: +{order_item.quantity}")
            
            serializer = OrderSerializer(order)
            return Response(serializer.data, status=status.HTTP_200_OK)
        
        except Order.DoesNotExist:
            return Response(
                {'error': 'Đơn hàng không tồn tại'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error updating order status: {str(e)}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Admin - Thống kê đơn hàng"""
        if not request.user.is_staff:
            return Response(
                {'error': 'Bạn không có quyền truy cập'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        total_orders = Order.objects.count()
        pending_orders = Order.objects.filter(status='pending').count()
        completed_orders = Order.objects.filter(status='delivered').count()
        total_revenue = sum(order.total_amount for order in Order.objects.all())
        
        return Response({
            'total_orders': total_orders,
            'pending_orders': pending_orders,
            'completed_orders': completed_orders,
            'total_revenue': float(total_revenue),
        })
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Xuất báo cáo Excel"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from datetime import datetime
        from django.db.models import Sum, Count, Q
        from products.models import Product
        
        # Get parameters
        report_type = request.GET.get('report_type', 'revenue')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        wb = Workbook()
        ws = wb.active
        
        header_fill = PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Filter orders by date range
        orders_query = Order.objects.all()
        if start_date:
            orders_query = orders_query.filter(created_at__gte=start_date)
        if end_date:
            orders_query = orders_query.filter(created_at__lte=end_date + ' 23:59:59')
        
        if report_type == 'revenue':
            ws.title = "Doanh Thu"
            ws.merge_cells('A1:F1')
            ws['A1'] = 'BÁO CÁO DOANH THU - WEB_TEDDY'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            # Date range info
            ws['A2'] = f'Từ ngày: {start_date if start_date else "Tất cả"} - Đến ngày: {end_date if end_date else "Tất cả"}'
            ws['A2'].alignment = Alignment(horizontal="center")
            
            headers = ['Mã ĐH', 'Khách hàng', 'Ngày', 'Tổng tiền', 'Trạng thái', 'Thanh toán']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            orders = orders_query.order_by('-created_at')
            total_revenue = 0
            for row, order in enumerate(orders, start=5):
                ws.cell(row=row, column=1, value=order.order_code).border = border
                ws.cell(row=row, column=2, value=order.full_name).border = border
                ws.cell(row=row, column=3, value=order.created_at.strftime('%d/%m/%Y')).border = border
                ws.cell(row=row, column=4, value=float(order.total_amount)).border = border
                ws.cell(row=row, column=5, value=order.get_status_display()).border = border
                ws.cell(row=row, column=6, value=order.get_payment_method_display()).border = border
                total_revenue += float(order.total_amount)
            
            # Summary row
            summary_row = len(orders) + 5
            ws.cell(row=summary_row, column=1, value='TỔNG CỘNG').font = Font(bold=True)
            ws.cell(row=summary_row, column=4, value=total_revenue).font = Font(bold=True)
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 20
        
        elif report_type == 'orders':
            ws.title = "Đơn Hàng"
            ws.merge_cells('A1:G1')
            ws['A1'] = 'BÁO CÁO ĐƠN HÀNG - WEB_TEDDY'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f'Từ ngày: {start_date if start_date else "Tất cả"} - Đến ngày: {end_date if end_date else "Tất cả"}'
            ws['A2'].alignment = Alignment(horizontal="center")
            
            headers = ['Mã ĐH', 'Khách hàng', 'SĐT', 'Địa chỉ', 'Tổng tiền', 'Trạng thái', 'Ngày']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            orders = orders_query.order_by('-created_at')
            for row, order in enumerate(orders, start=5):
                ws.cell(row=row, column=1, value=order.order_code).border = border
                ws.cell(row=row, column=2, value=order.full_name).border = border
                ws.cell(row=row, column=3, value=order.phone).border = border
                ws.cell(row=row, column=4, value=f"{order.address}, {order.district}, {order.city}").border = border
                ws.cell(row=row, column=5, value=float(order.total_amount)).border = border
                ws.cell(row=row, column=6, value=order.get_status_display()).border = border
                ws.cell(row=row, column=7, value=order.created_at.strftime('%d/%m/%Y')).border = border
            
            for col in ['A', 'B', 'C', 'E', 'F', 'G']:
                ws.column_dimensions[col].width = 20
            ws.column_dimensions['D'].width = 40
        
        elif report_type == 'products':
            ws.title = "Sản Phẩm Bán Chạy"
            ws.merge_cells('A1:F1')
            ws['A1'] = 'BÁO CÁO SẢN PHẨM BÁN CHẠY - WEB_TEDDY'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f'Từ ngày: {start_date if start_date else "Tất cả"} - Đến ngày: {end_date if end_date else "Tất cả"}'
            ws['A2'].alignment = Alignment(horizontal="center")
            
            headers = ['STT', 'Sản phẩm', 'Danh mục', 'Đã bán', 'Giá', 'Doanh thu']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            # Get order items in date range
            order_items_query = OrderItem.objects.filter(order__in=orders_query)
            product_stats = order_items_query.values('product__id', 'product__name', 'product__category__name').annotate(
                total_sold=Sum('quantity'),
                total_revenue=Sum('quantity') * Sum('product_price')
            ).order_by('-total_sold')
            
            for row, stat in enumerate(product_stats[:50], start=5):
                product = Product.objects.filter(id=stat['product__id']).first()
                ws.cell(row=row, column=1, value=row-4).border = border
                ws.cell(row=row, column=2, value=stat['product__name']).border = border
                ws.cell(row=row, column=3, value=stat['product__category__name'] or 'N/A').border = border
                ws.cell(row=row, column=4, value=stat['total_sold']).border = border
                ws.cell(row=row, column=5, value=float(product.price) if product else 0).border = border
                ws.cell(row=row, column=6, value=float(product.price * stat['total_sold']) if product else 0).border = border
            
            for col in ['A', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 18
            ws.column_dimensions['B'].width = 35
        
        elif report_type == 'customers':
            ws.title = "Khách Hàng"
            ws.merge_cells('A1:F1')
            ws['A1'] = 'BÁO CÁO KHÁCH HÀNG - WEB_TEDDY'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal="center")
            
            ws['A2'] = f'Từ ngày: {start_date if start_date else "Tất cả"} - Đến ngày: {end_date if end_date else "Tất cả"}'
            ws['A2'].alignment = Alignment(horizontal="center")
            
            headers = ['Khách hàng', 'Email', 'SĐT', 'Số đơn', 'Tổng chi tiêu', 'Ngày đầu']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            
            # Get customer statistics
            customer_stats = orders_query.values('user__username', 'email', 'phone', 'user__date_joined').annotate(
                order_count=Count('id'),
                total_spent=Sum('total_amount')
            ).order_by('-total_spent')
            
            for row, stat in enumerate(customer_stats, start=5):
                ws.cell(row=row, column=1, value=stat['user__username'] or 'Khách').border = border
                ws.cell(row=row, column=2, value=stat['email']).border = border
                ws.cell(row=row, column=3, value=stat['phone']).border = border
                ws.cell(row=row, column=4, value=stat['order_count']).border = border
                ws.cell(row=row, column=5, value=float(stat['total_spent'])).border = border
                ws.cell(row=row, column=6, value=stat['user__date_joined'].strftime('%d/%m/%Y') if stat['user__date_joined'] else 'N/A').border = border
            
            for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                ws.column_dimensions[col].width = 22
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=bao_cao_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        """Xuất báo cáo PDF"""
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from datetime import datetime
        from django.db.models import Sum, Count
        from products.models import Product
        import io
        
        # Get parameters
        report_type = request.GET.get('report_type', 'revenue')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()
        
        # Filter orders by date range
        orders_query = Order.objects.all()
        if start_date:
            orders_query = orders_query.filter(created_at__gte=start_date)
        if end_date:
            orders_query = orders_query.filter(created_at__lte=end_date + ' 23:59:59')
        
        if report_type == 'revenue':
            title = Paragraph('<para align=center><b>BAO CAO DOANH THU - WEB_TEDDY</b></para>', styles['Title'])
            elements.append(title)
            date_range = Paragraph(f'<para align=center>Tu ngay: {start_date if start_date else "Tat ca"} - Den ngay: {end_date if end_date else "Tat ca"}</para>', styles['Normal'])
            elements.append(date_range)
            elements.append(Spacer(1, 0.3*inch))
            
            orders = orders_query.order_by('-created_at')[:100]
            data = [['Ma DH', 'Khach hang', 'Ngay', 'Tong tien', 'Trang thai', 'Thanh toan']]
            
            total_revenue = 0
            for order in orders:
                data.append([
                    order.order_code,
                    order.full_name[:15],
                    order.created_at.strftime('%d/%m/%y'),
                    f'{order.total_amount:,.0f}',
                    order.get_status_display()[:10],
                    order.get_payment_method_display()[:10],
                ])
                total_revenue += float(order.total_amount)
            
            data.append(['', '', 'TONG CONG:', f'{total_revenue:,.0f}', '', ''])
            
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF69B4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -2), 7),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
        
        elif report_type == 'orders':
            title = Paragraph('<para align=center><b>BAO CAO DON HANG - WEB_TEDDY</b></para>', styles['Title'])
            elements.append(title)
            date_range = Paragraph(f'<para align=center>Tu ngay: {start_date if start_date else "Tat ca"} - Den ngay: {end_date if end_date else "Tat ca"}</para>', styles['Normal'])
            elements.append(date_range)
            elements.append(Spacer(1, 0.3*inch))
            
            orders = orders_query.order_by('-created_at')[:100]
            data = [['Ma DH', 'Khach', 'SDT', 'Tong tien', 'TT', 'Ngay']]
            
            for order in orders:
                data.append([
                    order.order_code,
                    order.full_name[:12],
                    order.phone,
                    f'{order.total_amount:,.0f}',
                    order.get_status_display()[:10],
                    order.created_at.strftime('%d/%m/%y')
                ])
            
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF69B4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
        
        elif report_type == 'products':
            title = Paragraph('<para align=center><b>BAO CAO SAN PHAM BAN CHAY - WEB_TEDDY</b></para>', styles['Title'])
            elements.append(title)
            date_range = Paragraph(f'<para align=center>Tu ngay: {start_date if start_date else "Tat ca"} - Den ngay: {end_date if end_date else "Tat ca"}</para>', styles['Normal'])
            elements.append(date_range)
            elements.append(Spacer(1, 0.3*inch))
            
            order_items_query = OrderItem.objects.filter(order__in=orders_query)
            product_stats = order_items_query.values('product__id', 'product__name', 'product__category__name').annotate(
                total_sold=Sum('quantity')
            ).order_by('-total_sold')[:50]
            
            data = [['STT', 'San pham', 'Danh muc', 'Da ban', 'Gia', 'Doanh thu']]
            
            for idx, stat in enumerate(product_stats, 1):
                product = Product.objects.filter(id=stat['product__id']).first()
                data.append([
                    str(idx),
                    stat['product__name'][:25],
                    (stat['product__category__name'] or 'N/A')[:15],
                    str(stat['total_sold']),
                    f'{product.price:,.0f}' if product else '0',
                    f'{(product.price * stat["total_sold"]):,.0f}' if product else '0'
                ])
            
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF69B4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
        
        elif report_type == 'customers':
            title = Paragraph('<para align=center><b>BAO CAO KHACH HANG - WEB_TEDDY</b></para>', styles['Title'])
            elements.append(title)
            date_range = Paragraph(f'<para align=center>Tu ngay: {start_date if start_date else "Tat ca"} - Den ngay: {end_date if end_date else "Tat ca"}</para>', styles['Normal'])
            elements.append(date_range)
            elements.append(Spacer(1, 0.3*inch))
            
            customer_stats = orders_query.values('user__username', 'email', 'phone').annotate(
                order_count=Count('id'),
                total_spent=Sum('total_amount')
            ).order_by('-total_spent')[:50]
            
            data = [['Khach hang', 'Email', 'SDT', 'So don', 'Tong chi tieu']]
            
            for stat in customer_stats:
                data.append([
                    (stat['user__username'] or 'Khach')[:15],
                    stat['email'][:20],
                    stat['phone'],
                    str(stat['order_count']),
                    f'{stat["total_spent"]:,.0f}'
                ])
            
            table = Table(data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF69B4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
        
        elements.append(table)
        doc.build(elements)
        
        pdf = buffer.getvalue()
        buffer.close()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=bao_cao_{report_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response.write(pdf)
        return response


@api_view(['POST'])
@permission_classes([AllowAny])
@csrf_exempt
def momo_callback(request):
    """
    Nhận callback từ MoMo sau khi thanh toán
    """
    try:
        data = request.data
        logger.info(f"MoMo callback received: {data}")
        
        # Xác thực signature
        if not MoMoPayment.verify_signature(data):
            logger.error("Invalid MoMo signature")
            return Response(
                {'message': 'Invalid signature'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Lấy thông tin từ callback
        result_code = data.get('resultCode')
        order_id = data.get('orderId')
        trans_id = data.get('transId')
        message = data.get('message')
        
        # Tìm order
        try:
            order = Order.objects.get(id=int(order_id))
        except (Order.DoesNotExist, ValueError):
            logger.error(f"Order not found: {order_id}")
            return Response(
                {'message': 'Order not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Cập nhật trạng thái thanh toán
        if result_code == 0:
            # Thanh toán thành công
            order.payment_status = 'completed'
            order.momo_transaction_id = trans_id
            order.save(update_fields=['payment_status', 'momo_transaction_id'])
            
            logger.info(f"Order {order.order_code} payment completed via MoMo")
            
            return Response({'message': 'Success'}, status=status.HTTP_200_OK)
        else:
            # Thanh toán thất bại
            order.payment_status = 'failed'
            order.save(update_fields=['payment_status'])
            
            logger.warning(f"Order {order.order_code} payment failed: {message}")
            
            return Response({'message': 'Payment failed'}, status=status.HTTP_200_OK)
    
    except Exception as e:
        logger.error(f"Error processing MoMo callback: {str(e)}")
        return Response(
            {'message': 'Internal error'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_momo_payment_status(request, order_id):
    """
    Kiểm tra trạng thái thanh toán MoMo của đơn hàng
    """
    try:
        # Lấy đơn hàng
        order = Order.objects.get(id=order_id, user=request.user)
        
        # Kiểm tra xem đơn hàng có dùng MoMo không
        if order.payment_method != 'momo':
            return Response(
                {'error': 'Đơn hàng không sử dụng thanh toán MoMo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Kiểm tra trạng thái từ MoMo
        if order.momo_request_id:
            momo_response = MoMoPayment.check_transaction_status(
                order_id=order.id,
                request_id=order.momo_request_id
            )
            
            if momo_response.get('resultCode') == 0:
                # Cập nhật trạng thái thanh toán
                order.payment_status = 'completed'
                order.momo_transaction_id = momo_response.get('transId')
                order.save(update_fields=['payment_status', 'momo_transaction_id'])
            
            return Response({
                'order': OrderSerializer(order).data,
                'momo_status': momo_response
            })
        else:
            return Response(
                {'error': 'Chưa có thông tin thanh toán MoMo'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    except Order.DoesNotExist:
        return Response(
            {'error': 'Đơn hàng không tồn tại'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error checking MoMo payment status: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
